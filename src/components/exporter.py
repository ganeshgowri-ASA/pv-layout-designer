"""
PV Layout Designer - Export & Reporting Module
Generates professional Excel BoQ and PDF reports with layout images
"""

from io import BytesIO, StringIO
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import ezdxf

# Constants
METERS_TO_DEGREES = 111000  # Approximate conversion factor: 1 degree ≈ 111 km at equator


def generate_excel_boq(layout: Dict[str, Any], config: Dict[str, Any]) -> BytesIO:
    """
    Generate Excel Bill of Quantities (BoQ) with multiple sheets
    
    Args:
        layout: Dictionary containing layout data with modules, rows, and metrics
        config: Configuration dictionary with project settings
        
    Returns:
        BytesIO: Excel file in memory
    """
    output = BytesIO()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SHEET 1: PROJECT SUMMARY =====
    ws_summary = wb.create_sheet("Project Summary")
    
    # Project Information
    summary_data = [
        ["PV LAYOUT DESIGNER - PROJECT SUMMARY", ""],
        ["", ""],
        ["Project Information", ""],
        ["Project Name", config.get("project_name", "Untitled Project")],
        ["Location", config.get("location", "Not Specified")],
        ["Date Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Designer", config.get("designer", "PV Layout Designer")],
        ["", ""],
        ["Site Specifications", ""],
        ["Site Area (m²)", layout.get("site_area", 0)],
        ["Usable Area (m²)", layout.get("usable_area", 0)],
        ["Latitude", config.get("latitude", "N/A")],
        ["Longitude", config.get("longitude", "N/A")],
        ["", ""],
        ["Layout Summary", ""],
        ["Total Modules", layout.get("total_modules", 0)],
        ["Total Capacity (kWp)", layout.get("total_capacity_kwp", 0)],
        ["Total Capacity (MWp)", layout.get("total_capacity_kwp", 0) / 1000],
        ["Number of Rows", layout.get("num_rows", 0)],
        ["", ""],
        ["Technical Parameters", ""],
        ["Ground Coverage Ratio (GCR)", f"{layout.get('gcr', 0):.2%}"],
        ["Module Tilt Angle (°)", config.get("tilt_angle", 0)],
        ["Module Orientation", config.get("orientation", "Portrait")],
        ["Row Orientation", config.get("row_orientation", "North-South")],
        ["Inter-Row Spacing (m)", layout.get("inter_row_spacing", 0)],
        ["Module Height (m)", config.get("module_height", 0)],
        ["", ""],
        ["Module Specifications", ""],
        ["Module Power (Wp)", config.get("module_power", 550)],
        ["Module Length (m)", config.get("module_length", 2.278)],
        ["Module Width (m)", config.get("module_width", 1.134)],
        ["Modules per Structure", config.get("modules_per_structure", 28)],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Style title
            if row_idx == 1:
                cell.font = Font(color="FFFFFF", bold=True, size=14)
                cell.fill = header_fill
                
            # Style section headers
            elif col_idx == 1 and value and ":" not in str(value) and row_data[1] == "":
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Merge title cells
    ws_summary.merge_cells('A1:B1')
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    
    # ===== SHEET 2: MODULE LIST =====
    ws_modules = wb.create_sheet("Module List")
    
    # Generate sample module data
    modules = layout.get("modules", [])
    if not modules and layout.get("total_modules", 0) > 0:
        # Generate sample modules if not provided
        total_modules = layout.get("total_modules", 100)
        num_rows = layout.get("num_rows", 10)
        modules_per_row = total_modules // num_rows
        
        modules = []
        for row_num in range(1, num_rows + 1):
            for mod_num in range(1, modules_per_row + 1):
                modules.append({
                    "module_id": (row_num - 1) * modules_per_row + mod_num,
                    "row": row_num,
                    "position": mod_num,
                    "latitude": config.get("latitude", 23.0) + (row_num * 0.0001),
                    "longitude": config.get("longitude", 72.5) + (mod_num * 0.0001),
                    "status": "Active"
                })
    
    module_headers = ["Module #", "Row", "Position in Row", "Latitude", "Longitude", "Status"]
    ws_modules.append(module_headers)
    
    # Style headers
    for col_idx, header in enumerate(module_headers, 1):
        cell = ws_modules.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Add module data
    for module in modules[:1000]:  # Limit to 1000 modules for Excel
        row_data = [
            module.get("module_id", ""),
            module.get("row", ""),
            module.get("position", ""),
            f"{module.get('latitude', 0):.6f}" if module.get('latitude') else "N/A",
            f"{module.get('longitude', 0):.6f}" if module.get('longitude') else "N/A",
            module.get("status", "Active")
        ]
        ws_modules.append(row_data)
        
        # Apply borders
        for col_idx in range(1, len(module_headers) + 1):
            ws_modules.cell(row=ws_modules.max_row, column=col_idx).border = border
    
    # Auto-size columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_modules.column_dimensions[col].width = 15
    
    # ===== SHEET 3: BILL OF QUANTITIES =====
    ws_boq = wb.create_sheet("Bill of Quantities")
    
    # Calculate quantities
    total_modules = layout.get("total_modules", 0)
    modules_per_structure = config.get("modules_per_structure", 28)
    num_structures = int(total_modules / modules_per_structure) if modules_per_structure > 0 else 0
    
    # Cable length estimation (rough estimate)
    num_rows = layout.get("num_rows", 0)
    avg_row_length = 100  # meters, estimated
    dc_cable_length = num_rows * avg_row_length * 1.2  # 20% contingency
    
    # Number of inverters (assume 1 inverter per 500kW)
    total_capacity_kw = layout.get("total_capacity_kwp", 0)
    num_inverters = max(1, int(total_capacity_kw / 500))
    
    boq_data = [
        ["Category", "Item", "Specification", "Quantity", "Unit", "Remarks"],
        ["Modules", "PV Module", f"{config.get('module_power', 550)}Wp Monocrystalline", total_modules, "Nos", "Including frames and junction boxes"],
        ["Structure", "Mounting Structure", f"{modules_per_structure} modules per structure", num_structures, "Sets", "Hot-dip galvanized steel"],
        ["Structure", "Foundation", "Concrete/Pile foundation", num_structures, "Sets", "As per soil condition"],
        ["Cables", "DC Cable", "4mm² Solar cable", int(dc_cable_length), "Meters", "UV resistant, -40°C to +90°C"],
        ["Cables", "AC Cable", "3C x 240mm² XLPE", int(num_inverters * 100), "Meters", "Underground armored cable"],
        ["Equipment", "String Inverter", "500kW Central inverter", num_inverters, "Nos", "IP65 rated, outdoor type"],
        ["Equipment", "Combiner Box", "16 String inputs", int(num_structures / 10), "Nos", "With SPD and DC breakers"],
        ["Equipment", "Transformer", "1MVA, 33/0.4kV", max(1, int(total_capacity_kw / 1000)), "Nos", "Outdoor type with OLTC"],
        ["Protection", "Lightning Arrestor", "Type 1+2 SPD", int(num_structures / 20), "Nos", "DC and AC side protection"],
        ["Protection", "Earthing System", "Complete earthing", 1, "Lot", "As per IS standards"],
        ["Civil", "Access Roads", "Compacted gravel roads", 1, "Lot", "3m width"],
        ["Civil", "Perimeter Fencing", "Chain link fencing", int(layout.get("site_area", 10000) ** 0.5 * 4), "Meters", "2.4m height with barbed wire"],
        ["Monitoring", "SCADA System", "Complete monitoring", 1, "Set", "Web-based with mobile app"],
        ["Monitoring", "Weather Station", "Irradiance & temperature", 1, "Set", "With data logger"],
    ]
    
    # Write BoQ data
    for row_data in boq_data:
        ws_boq.append(row_data)
        
        # Apply styles
        row_idx = ws_boq.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_boq.cell(row=row_idx, column=col_idx)
            cell.border = border
            
            # Header row
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-size columns
    ws_boq.column_dimensions['A'].width = 15
    ws_boq.column_dimensions['B'].width = 25
    ws_boq.column_dimensions['C'].width = 30
    ws_boq.column_dimensions['D'].width = 12
    ws_boq.column_dimensions['E'].width = 10
    ws_boq.column_dimensions['F'].width = 35
    
    # Save workbook to BytesIO
    wb.save(output)
    output.seek(0)
    
    return output


def generate_pdf_report(layout: Dict[str, Any], config: Dict[str, Any], 
                       images: Optional[Dict[str, Any]] = None) -> BytesIO:
    """
    Generate comprehensive PDF report with layout images and specifications
    
    Args:
        layout: Dictionary containing layout data
        config: Configuration dictionary with project settings
        images: Optional dictionary containing layout visualization images
        
    Returns:
        BytesIO: PDF file in memory
    """
    output = BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )
    
    # Container for PDF elements
    story = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    # ===== COVER PAGE =====
    story.append(Spacer(1, 2*inch))
    
    title = Paragraph("PV LAYOUT DESIGNER", title_style)
    story.append(title)
    story.append(Spacer(1, 0.3*inch))
    
    subtitle = Paragraph("Solar PV Plant Layout Report", styles['Heading2'])
    subtitle.alignment = TA_CENTER
    story.append(subtitle)
    story.append(Spacer(1, 1*inch))
    
    # Project details table
    project_data = [
        ['Project Name:', config.get('project_name', 'Untitled Project')],
        ['Location:', config.get('location', 'Not Specified')],
        ['Date:', datetime.now().strftime("%B %d, %Y")],
        ['Total Capacity:', f"{layout.get('total_capacity_kwp', 0) / 1000:.2f} MWp"],
    ]
    
    project_table = Table(project_data, colWidths=[2*inch, 4*inch])
    project_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 12),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 12),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#366092')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(project_table)
    
    story.append(PageBreak())
    
    # ===== LAYOUT SPECIFICATIONS =====
    story.append(Paragraph("Layout Specifications", heading_style))
    story.append(Spacer(1, 0.2*inch))
    
    specs_data = [
        ['Parameter', 'Value', 'Unit'],
        ['Site Area', f"{layout.get('site_area', 0):,.0f}", 'm²'],
        ['Usable Area', f"{layout.get('usable_area', 0):,.0f}", 'm²'],
        ['Total Modules', f"{layout.get('total_modules', 0):,}", 'Nos'],
        ['Total Capacity', f"{layout.get('total_capacity_kwp', 0):,.2f}", 'kWp'],
        ['Number of Rows', str(layout.get('num_rows', 0)), '-'],
        ['Ground Coverage Ratio', f"{layout.get('gcr', 0):.1%}", '-'],
        ['Module Power', str(config.get('module_power', 550)), 'Wp'],
        ['Module Tilt Angle', str(config.get('tilt_angle', 25)), '°'],
        ['Module Orientation', config.get('orientation', 'Portrait'), '-'],
        ['Inter-Row Spacing', f"{layout.get('inter_row_spacing', 0):.2f}", 'm'],
    ]
    
    specs_table = Table(specs_data, colWidths=[3*inch, 2*inch, 1*inch])
    specs_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(specs_table)
    story.append(Spacer(1, 0.4*inch))
    
    # ===== LAYOUT IMAGES =====
    if images:
        story.append(Paragraph("Layout Visualizations", heading_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Add images if provided
        for img_name, img_data in images.items():
            try:
                if isinstance(img_data, (str, BytesIO)):
                    img = Image(img_data, width=6*inch, height=4*inch)
                    story.append(Paragraph(img_name, styles['Heading3']))
                    story.append(img)
                    story.append(Spacer(1, 0.3*inch))
            except (IOError, OSError, ValueError) as e:
                # Skip if image can't be loaded (file not found, invalid format, etc.)
                story.append(Paragraph(f"[Image: {img_name} - Could not load]", styles['Normal']))
                story.append(Spacer(1, 0.2*inch))
    
    # ===== EQUIPMENT SUMMARY =====
    story.append(PageBreak())
    story.append(Paragraph("Equipment Summary", heading_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Calculate equipment quantities
    total_modules = layout.get("total_modules", 0)
    modules_per_structure = config.get("modules_per_structure", 28)
    num_structures = int(total_modules / modules_per_structure) if modules_per_structure > 0 else 0
    total_capacity_kw = layout.get("total_capacity_kwp", 0)
    num_inverters = max(1, int(total_capacity_kw / 500))
    
    equipment_data = [
        ['Equipment', 'Quantity', 'Specification'],
        ['PV Modules', f"{total_modules:,}", f"{config.get('module_power', 550)}Wp Monocrystalline"],
        ['Mounting Structures', f"{num_structures:,}", f"{modules_per_structure} modules/structure"],
        ['String Inverters', str(num_inverters), "500kW rated"],
        ['Combiner Boxes', str(int(num_structures / 10)), "16 string inputs"],
        ['Transformers', str(max(1, int(total_capacity_kw / 1000))), "1MVA, 33/0.4kV"],
    ]
    
    equipment_table = Table(equipment_data, colWidths=[3*inch, 1.5*inch, 2.5*inch])
    equipment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    story.append(equipment_table)
    
    # ===== FOOTER NOTE =====
    story.append(Spacer(1, 0.5*inch))
    footer_text = f"Report generated by PV Layout Designer on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    story.append(Paragraph(footer_text, styles['Normal']))
    
    # Build PDF
    doc.build(story)
    output.seek(0)
    
    return output


def generate_dxf_export(layout: Dict[str, Any]) -> BytesIO:
    """
    Generate DXF file for CAD software import (optional feature)
    
    Args:
        layout: Dictionary containing layout data with module positions
        
    Returns:
        BytesIO: DXF file in memory
    """
    # Create new DXF document
    doc = ezdxf.new('R2010')  # AutoCAD 2010 format
    msp = doc.modelspace()
    
    # Define layers
    doc.layers.new(name='MODULES', dxfattribs={'color': 3})  # Green
    doc.layers.new(name='STRUCTURES', dxfattribs={'color': 5})  # Blue
    doc.layers.new(name='SITE_BOUNDARY', dxfattribs={'color': 1})  # Red
    doc.layers.new(name='DIMENSIONS', dxfattribs={'color': 7})  # White/Black
    
    # Draw site boundary
    boundary = layout.get('site_boundary', [])
    if boundary and len(boundary) >= 3:
        points = [(p.get('lon', 0), p.get('lat', 0)) for p in boundary]
        points.append(points[0])  # Close the polygon
        msp.add_lwpolyline(points, dxfattribs={'layer': 'SITE_BOUNDARY'})
    
    # Draw modules
    modules = layout.get('modules', [])
    module_length = layout.get('module_length', 2.278)
    module_width = layout.get('module_width', 1.134)
    
    for module in modules:
        x = module.get('longitude', 0)
        y = module.get('latitude', 0)
        
        # Draw module as rectangle
        points = [
            (x, y),
            (x + module_length/METERS_TO_DEGREES, y),  # Convert meters to degrees
            (x + module_length/METERS_TO_DEGREES, y + module_width/METERS_TO_DEGREES),
            (x, y + module_width/METERS_TO_DEGREES),
            (x, y)
        ]
        msp.add_lwpolyline(points, dxfattribs={'layer': 'MODULES'})
        
        # Add module ID text
        module_id = module.get('module_id', '')
        if module_id:
            text_x = x + module_length/(METERS_TO_DEGREES * 2)
            text_y = y + module_width/(METERS_TO_DEGREES * 2)
            msp.add_text(
                str(module_id),
                dxfattribs={
                    'layer': 'MODULES',
                    'height': 0.5,
                    'insert': (text_x, text_y, 0)
                }
            )
    
    # Add title block
    title_text = f"PV Layout - {layout.get('total_modules', 0)} Modules"
    msp.add_text(
        title_text,
        dxfattribs={
            'layer': 'DIMENSIONS',
            'height': 2,
            'insert': (0, 0, 0)
        }
    )
    
    # Write to StringIO first, then convert to BytesIO
    string_output = StringIO()
    doc.write(string_output)
    
    # Convert string to bytes
    output = BytesIO(string_output.getvalue().encode('utf-8'))
    output.seek(0)
    
    return output
