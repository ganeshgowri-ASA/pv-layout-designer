"""
Unit tests for the exporter module
Tests Excel BoQ, PDF report, and DXF export generation
"""

import pytest
from io import BytesIO
from src.components.exporter import generate_excel_boq, generate_pdf_report, generate_dxf_export
import openpyxl
from reportlab.lib.pagesizes import A4


# Sample test data
@pytest.fixture
def sample_layout():
    """Sample layout data for testing"""
    return {
        'site_area': 50000,
        'usable_area': 45000,
        'total_modules': 5000,
        'total_capacity_kwp': 2750,
        'num_rows': 178,
        'gcr': 0.35,
        'inter_row_spacing': 4.5,
        'module_length': 2.278,
        'module_width': 1.134,
        'modules': [
            {
                'module_id': i,
                'row': (i // 28) + 1,
                'position': (i % 28) + 1,
                'latitude': 23.0225 + (i * 0.0001),
                'longitude': 72.5714 + (i * 0.0001),
                'status': 'Active'
            }
            for i in range(100)
        ],
        'site_boundary': [
            {'lat': 23.0, 'lon': 72.5},
            {'lat': 23.0, 'lon': 72.6},
            {'lat': 23.1, 'lon': 72.6},
            {'lat': 23.1, 'lon': 72.5},
        ]
    }


@pytest.fixture
def sample_config():
    """Sample configuration for testing"""
    return {
        'project_name': 'Test Solar Plant',
        'location': 'Gujarat, India',
        'latitude': 23.0225,
        'longitude': 72.5714,
        'designer': 'Test Designer',
        'module_power': 550,
        'module_length': 2.278,
        'module_width': 1.134,
        'tilt_angle': 25,
        'orientation': 'Portrait',
        'row_orientation': 'North-South',
        'module_height': 1.5,
        'modules_per_structure': 28
    }


class TestExcelBoQ:
    """Test Excel BoQ generation"""
    
    def test_generate_excel_returns_bytesio(self, sample_layout, sample_config):
        """Test that generate_excel_boq returns BytesIO object"""
        result = generate_excel_boq(sample_layout, sample_config)
        assert isinstance(result, BytesIO)
        assert result.tell() == 0  # Should be at start of stream
    
    def test_excel_has_correct_sheets(self, sample_layout, sample_config):
        """Test that Excel file has all required sheets"""
        excel_file = generate_excel_boq(sample_layout, sample_config)
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_file)
        
        # Check sheet names
        sheet_names = wb.sheetnames
        assert 'Project Summary' in sheet_names
        assert 'Module List' in sheet_names
        assert 'Bill of Quantities' in sheet_names
    
    def test_excel_project_summary_data(self, sample_layout, sample_config):
        """Test that project summary contains correct data"""
        excel_file = generate_excel_boq(sample_layout, sample_config)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Project Summary']
        
        # Check for key values
        found_project_name = False
        found_location = False
        
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'Project Name':
                assert row[1] == 'Test Solar Plant'
                found_project_name = True
            if row[0] == 'Location':
                assert row[1] == 'Gujarat, India'
                found_location = True
        
        assert found_project_name, "Project name not found in summary"
        assert found_location, "Location not found in summary"
    
    def test_excel_module_list_headers(self, sample_layout, sample_config):
        """Test that module list has correct headers"""
        excel_file = generate_excel_boq(sample_layout, sample_config)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Module List']
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert 'Module #' in headers
        assert 'Row' in headers
        assert 'Position in Row' in headers
        assert 'Latitude' in headers
        assert 'Longitude' in headers
        assert 'Status' in headers
    
    def test_excel_boq_categories(self, sample_layout, sample_config):
        """Test that BoQ sheet has correct categories"""
        excel_file = generate_excel_boq(sample_layout, sample_config)
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Bill of Quantities']
        
        # Collect all categories
        categories = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                categories.add(row[0])
        
        # Check for expected categories
        assert 'Modules' in categories
        assert 'Structure' in categories
        assert 'Cables' in categories
        assert 'Equipment' in categories


class TestPDFReport:
    """Test PDF report generation"""
    
    def test_generate_pdf_returns_bytesio(self, sample_layout, sample_config):
        """Test that generate_pdf_report returns BytesIO object"""
        result = generate_pdf_report(sample_layout, sample_config)
        assert isinstance(result, BytesIO)
        assert result.tell() == 0  # Should be at start of stream
    
    def test_pdf_has_content(self, sample_layout, sample_config):
        """Test that PDF file has content"""
        pdf_file = generate_pdf_report(sample_layout, sample_config)
        content = pdf_file.read()
        
        # PDF should start with %PDF
        assert content.startswith(b'%PDF')
        
        # Should have reasonable size (> 1KB)
        assert len(content) > 1024
    
    def test_pdf_with_images(self, sample_layout, sample_config):
        """Test PDF generation with images parameter"""
        images = {
            'Top View': None,  # Can be image path or BytesIO
            'Side View': None
        }
        
        pdf_file = generate_pdf_report(sample_layout, sample_config, images)
        assert isinstance(pdf_file, BytesIO)
        content = pdf_file.read()
        assert len(content) > 0


class TestDXFExport:
    """Test DXF export generation"""
    
    def test_generate_dxf_returns_bytesio(self, sample_layout):
        """Test that generate_dxf_export returns BytesIO object"""
        result = generate_dxf_export(sample_layout)
        assert isinstance(result, BytesIO)
        assert result.tell() == 0  # Should be at start of stream
    
    def test_dxf_has_content(self, sample_layout):
        """Test that DXF file has content"""
        dxf_file = generate_dxf_export(sample_layout)
        content = dxf_file.read()
        
        # DXF files should have specific markers
        content_str = content.decode('utf-8', errors='ignore')
        assert 'SECTION' in content_str or 'HEADER' in content_str
        
        # Should have reasonable size
        assert len(content) > 100
    
    def test_dxf_with_empty_layout(self):
        """Test DXF generation with minimal layout data"""
        minimal_layout = {
            'total_modules': 0,
            'modules': [],
            'site_boundary': []
        }
        
        dxf_file = generate_dxf_export(minimal_layout)
        assert isinstance(dxf_file, BytesIO)


class TestEdgeCases:
    """Test edge cases and error handling"""
    
    def test_empty_layout_excel(self, sample_config):
        """Test Excel generation with empty layout"""
        empty_layout = {
            'site_area': 0,
            'usable_area': 0,
            'total_modules': 0,
            'total_capacity_kwp': 0,
            'num_rows': 0,
            'gcr': 0,
            'inter_row_spacing': 0,
            'modules': []
        }
        
        excel_file = generate_excel_boq(empty_layout, sample_config)
        assert isinstance(excel_file, BytesIO)
    
    def test_empty_layout_pdf(self, sample_config):
        """Test PDF generation with empty layout"""
        empty_layout = {
            'site_area': 0,
            'usable_area': 0,
            'total_modules': 0,
            'total_capacity_kwp': 0,
            'num_rows': 0,
            'gcr': 0,
            'inter_row_spacing': 0,
            'modules': []
        }
        
        pdf_file = generate_pdf_report(empty_layout, sample_config)
        assert isinstance(pdf_file, BytesIO)
    
    def test_missing_optional_fields(self):
        """Test generation with missing optional fields"""
        minimal_layout = {'total_modules': 100}
        minimal_config = {'project_name': 'Test'}
        
        # Should not raise exceptions
        excel_file = generate_excel_boq(minimal_layout, minimal_config)
        pdf_file = generate_pdf_report(minimal_layout, minimal_config)
        
        assert isinstance(excel_file, BytesIO)
        assert isinstance(pdf_file, BytesIO)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
